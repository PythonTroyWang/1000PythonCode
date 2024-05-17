import openpyxl
from xml_util import str_to_date_and_get_month


def output_excel_file(data, company_name, output_path):
    workbook = openpyxl.Workbook()
    ws_total = workbook.active
    ws_total.title = '统计'

    ws_total.append(['月份', '进项金额', '', '', '销项金额', '', ''])
    ws_total.merge_cells("B1:D1")
    ws_total.merge_cells("E1:G1")
    ws_total.merge_cells("A1:A2")

    ws_total.append(['月份', '金额', '税额', '总金额', '金额', '税额', '总金额'])

    input_amount_calc = 0
    input_tax_amount_calc = 0
    input_total_amount_calc = 0
    output_amount_calc = 0
    output_tax_amount_calc = 0
    output_total_amount_calc = 0

    for i in range(1, 13):
        input_amount = 0
        input_tax_amount = 0
        input_total_amount = 0
        output_amount = 0
        output_tax_amount = 0
        output_total_amount = 0
        for item in data:
            month = int(str_to_date_and_get_month(item[0]))
            if i == month:
                seller_name = item[2]
                buyer_name = item[3]
                amount = item[4]
                tax_total = item[5]
                total_amount = item[6]
                if buyer_name == company_name:
                    input_amount += float(amount)
                    input_tax_amount += float(tax_total)
                    input_total_amount += float(total_amount)
                elif seller_name == company_name:
                    output_amount += float(amount)
                    output_tax_amount += float(tax_total)
                    output_total_amount += float(total_amount)
        ws_total.append([str(i) + '月', input_amount, input_tax_amount, input_total_amount, output_amount,
                         output_tax_amount, output_total_amount])
        input_amount_calc += input_amount
        input_tax_amount_calc += input_tax_amount
        input_total_amount_calc += input_total_amount
        output_amount_calc += output_amount
        output_tax_amount_calc += output_tax_amount
        output_total_amount_calc += output_total_amount

    ws_total.append(['合计', input_amount_calc, input_tax_amount_calc, input_total_amount_calc, output_amount_calc,
                     output_tax_amount_calc, output_total_amount_calc])

    ws_input_detail = workbook.create_sheet('进项明细表', 1)
    ws_input_detail.append(['发票号码', '所属月份', '销售方名称', '购买方名称', '金额', '税额', '总金额', '发票状态', '发票类型'])

    ws_output_detail = workbook.create_sheet('销项明细表', 2)
    ws_output_detail.append(['发票号码', '所属月份', '销售方名称', '购买方名称', '金额', '税额', '总金额', '发票状态', '发票类型'])

    add_detail(company_name, data, ws_input_detail, ws_output_detail)

    workbook.save(output_path + f'/{company_name}数电发票数据统计.xlsx')


def add_detail(company_name, data, ws_detail, ws_output_detail):
    for item in data:
        invoice_number = item[1]
        month = str_to_date_and_get_month(item[0])
        seller_name = item[2]
        buyer_name = item[3]
        amount = item[4]
        tax_total = item[5]
        total_amount = item[6]
        invoice_status = item[7]
        invoice_type = item[8]
        if invoice_status == "Y":
            invoice_status = "正常"
        else:
            invoice_status = "红冲"
        if buyer_name == company_name:
            ws_detail.append([invoice_number, str(month) + '月', seller_name, buyer_name, amount, tax_total, total_amount, invoice_status, invoice_type])
        elif seller_name == company_name:
            ws_output_detail.append([invoice_number, str(month) + '月', seller_name, buyer_name, amount, tax_total, total_amount, invoice_status, invoice_type])
        else:
            print("名称和发票的名称不符")
